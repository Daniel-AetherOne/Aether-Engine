from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None


CANONICAL_FILES = [
    ("articles", "articles.csv"),
    ("tiers", "tiers.csv"),
    ("supplier_factors", "supplier_factors.csv"),
    ("transport", "transport.csv"),
    ("customers", "customers.xlsx"),
]


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def count_rows_csv(path: Path, *, delimiter: str = ";") -> int:
    """
    Counts data rows (excludes header). Skips empty lines.
    Strict-ish: assumes first non-empty row is header.
    """
    import csv

    if not path.exists():
        return 0

    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)
        header_found = False
        count = 0
        for row in reader:
            if not row or all((c is None or str(c).strip() == "") for c in row):
                continue
            if not header_found:
                header_found = True
                continue
            count += 1
        return count


def count_rows_xlsx(path: Path) -> int:
    """
    Counts data rows (excludes header). Skips empty rows.
    Uses read_only=True; MUST close workbook to avoid Windows locks.
    """
    if load_workbook is None:
        return 0
    if not path.exists():
        return 0

    wb = None
    try:
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
        ws = wb.active

        header_found = False
        count = 0
        for row in ws.iter_rows(values_only=True):
            if row is None or all((c is None or str(c).strip() == "") for c in row):
                continue
            if not header_found:
                header_found = True
                continue
            count += 1
        return count
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


@dataclass(frozen=True)
class ManifestDatasetEntry:
    type: str
    versionId: str
    filename: str
    uploadedBy: str
    uploadedAt: str
    checksum: str
    rowCount: int


@dataclass(frozen=True)
class ActiveManifest:
    activeVersionId: str
    datasets: list[ManifestDatasetEntry]

    def to_dict(self) -> dict[str, Any]:
        return {
            "activeVersionId": self.activeVersionId,
            "datasets": [asdict(d) for d in self.datasets],
        }


def build_manifest_for_dataset_dir(
    dataset_dir: Path,
    *,
    version_id: str,
    uploaded_by: str,
    uploaded_at: Optional[str] = None,
) -> ActiveManifest:
    """
    Builds a manifest.json for a dataset directory (active or staging snapshot).
    Deterministic: checksums and rowCounts are derived from file bytes/content.
    """
    uploaded_at = uploaded_at or utc_now_iso()

    entries: list[ManifestDatasetEntry] = []
    for dtype, fname in CANONICAL_FILES:
        p = dataset_dir / fname
        checksum = sha256_file(p) if p.exists() else ""
        if fname.endswith(".csv"):
            row_count = count_rows_csv(p)
        else:
            row_count = count_rows_xlsx(p)

        entries.append(
            ManifestDatasetEntry(
                type=dtype,
                versionId=version_id,
                filename=fname,
                uploadedBy=uploaded_by,
                uploadedAt=uploaded_at,
                checksum=checksum,
                rowCount=row_count,
            )
        )

    return ActiveManifest(activeVersionId=version_id, datasets=entries)


def read_manifest(dataset_dir: Path) -> dict[str, Any]:
    p = dataset_dir / "manifest.json"
    if not p.exists():
        raise FileNotFoundError(f"manifest.json not found in {dataset_dir}")
    with p.open("r", encoding="utf-8") as f:
        return json.load(f)


def write_manifest(dataset_dir: Path, manifest: dict[str, Any]) -> None:
    dataset_dir.mkdir(parents=True, exist_ok=True)
    p = dataset_dir / "manifest.json"
    tmp = dataset_dir / ".manifest.json.tmp"
    with tmp.open("w", encoding="utf-8", newline="\n") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)
        f.write("\n")
    tmp.replace(p)
