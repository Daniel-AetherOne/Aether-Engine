from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Optional

import csv

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None


# =========================
# 2.0 legacy (optioneel; kan later weg)
# =========================


@dataclass(frozen=True)
class ValidationIssue:
    code: str
    message: str
    path: str


# =========================
# 2.2 Parsing config (MVP: strict)
# =========================

CSV_DELIMITER = ";"
DECIMAL_SEPARATOR = ","


class ParseError(ValueError):
    """Raised when a file cannot be parsed deterministically according to our strict rules."""


def _trim(v: Any) -> Any:
    if isinstance(v, str):
        return v.strip()
    return v


def _is_row_empty(values: Iterable[Any]) -> bool:
    for v in values:
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        return False
    return True


def require_headers(headers: list[str], required: list[str], *, source: str) -> None:
    """
    Fail fast if required headers are missing.
    Comparison is case-insensitive + trims whitespace.
    """
    normalized = {h.strip().lower() for h in headers if h is not None}
    missing = [h for h in required if h.strip().lower() not in normalized]
    if missing:
        raise ParseError(f"{source}: missing required headers: {missing}")


def parse_csv(
    file_path: Path,
    *,
    required_headers: Optional[list[str]] = None,
    delimiter: str = CSV_DELIMITER,
) -> list[dict[str, Any]]:
    """
    Strict CSV parsing:
    - delimiter fixed (default ';')
    - trims whitespace on headers + cells
    - skips empty lines
    - headers required (first non-empty row)
    - returns List[dict]
    """
    if not file_path.exists():
        raise ParseError(f"CSV not found: {file_path}")

    rows: list[dict[str, Any]] = []

    with file_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)

        # find first non-empty row as header
        header_row: Optional[list[str]] = None
        for raw in reader:
            if _is_row_empty(raw):
                continue
            header_row = [str(_trim(c)) for c in raw]
            break

        if header_row is None:
            raise ParseError(f"CSV empty/no header: {file_path}")

        headers = [h.strip() for h in header_row]
        if any(h == "" for h in headers):
            raise ParseError(f"CSV has empty header names: {file_path}")

        if required_headers:
            require_headers(headers, required_headers, source=str(file_path))

        # read data rows
        for raw in reader:
            if _is_row_empty(raw):
                continue

            # strict: row length must match header length
            if len(raw) != len(headers):
                raise ParseError(
                    f"CSV row has {len(raw)} cols but header has {len(headers)} cols: {file_path}"
                )

            item: dict[str, Any] = {}
            for i, h in enumerate(headers):
                item[h] = _trim(raw[i])
            rows.append(item)

    return rows


def parse_xlsx(
    file_path: Path,
    *,
    sheet_name: Optional[str] = None,
    required_headers: Optional[list[str]] = None,
) -> list[dict[str, Any]]:
    """
    Strict XLSX parsing:
    - reads a sheet (by name or active sheet)
    - trims whitespace on headers + string cells
    - skips empty rows
    - headers required (first non-empty row)
    - returns List[dict]
    """
    if load_workbook is None:
        raise ParseError("openpyxl not installed; cannot parse .xlsx")

    if not file_path.exists():
        raise ParseError(f"XLSX not found: {file_path}")

    wb = None
    try:
        try:
            wb = load_workbook(filename=str(file_path), data_only=True, read_only=True)
        except Exception as e:
            raise ParseError(f"XLSX unreadable (not a valid .xlsx): {file_path}") from e

        ws = wb[sheet_name] if sheet_name else wb.active

        # find header
        header_row: Optional[list[str]] = None
        for row in ws.iter_rows(values_only=True):
            if row is None or _is_row_empty(row):
                continue
            header_row = [str(_trim(c)) if c is not None else "" for c in row]
            break

        if header_row is None:
            raise ParseError(f"XLSX empty/no header: {file_path}")

        headers = [h.strip() for h in header_row]
        if any(h == "" for h in headers):
            raise ParseError(f"XLSX has empty header names: {file_path}")

        if required_headers:
            require_headers(headers, required_headers, source=str(file_path))

        # read data rows after header
        rows: list[dict[str, Any]] = []
        started = False

        for row in ws.iter_rows(values_only=True):
            if row is None or _is_row_empty(row):
                if started:
                    continue
                else:
                    continue

            if not started:
                # first non-empty row was header; mark started and skip it
                started = True
                continue

            if len(row) != len(headers):
                raise ParseError(
                    f"XLSX row has {len(row)} cols but header has {len(headers)} cols: {file_path}"
                )

            item: dict[str, Any] = {}
            for i, h in enumerate(headers):
                item[h] = _trim(row[i])
            rows.append(item)

        return rows

    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def parse_decimal(value: Any, *, source: str) -> float:
    """
    Deterministic decimal parsing for NL format:
    - decimal separator = ','
    - no thousands separators in MVP
    """
    if value is None:
        raise ParseError(f"{source}: value is null")

    if isinstance(value, (int, float)):
        return float(value)

    if not isinstance(value, str):
        raise ParseError(f"{source}: expected str/number, got {type(value).__name__}")

    s = value.strip()
    if s == "":
        raise ParseError(f"{source}: empty number")

    # strict: no thousands separators in MVP
    if "." in s:
        raise ParseError(f"{source}: '.' not allowed as decimal separator (use ',')")

    s2 = s.replace(DECIMAL_SEPARATOR, ".")
    try:
        return float(s2)
    except Exception:
        raise ParseError(f"{source}: invalid decimal: {value!r}")


# =========================
# 2.3 Validation outputs + helpers
# =========================


@dataclass(frozen=True)
class ValidationError:
    datasetType: str
    rowNumber: Optional[int]  # None = file-level error
    field: Optional[str]
    errorCode: str
    message: str


@dataclass(frozen=True)
class ValidationWarning:
    datasetType: str
    rowNumber: Optional[int]
    field: Optional[str]
    warningCode: str
    message: str


@dataclass
class ValidationResult:
    ok: bool
    errors: list[ValidationError]
    warnings: list[ValidationWarning]


def _err(
    datasetType: str,
    rowNumber: Optional[int],
    field: Optional[str],
    errorCode: str,
    message: str,
) -> ValidationError:
    return ValidationError(datasetType, rowNumber, field, errorCode, message)


def _warn(
    datasetType: str,
    rowNumber: Optional[int],
    field: Optional[str],
    warningCode: str,
    message: str,
) -> ValidationWarning:
    return ValidationWarning(datasetType, rowNumber, field, warningCode, message)


def normalize_header(h: str) -> str:
    return (h or "").strip().lower()


def get_cell(row: dict[str, Any], header_aliases: list[str]) -> Any:
    """
    Header lookup case-insensitive, supports aliases.
    Uses the original keys as provided by parse_csv/parse_xlsx.
    """
    norm_map = {normalize_header(k): k for k in row.keys()}
    for alias in header_aliases:
        k = norm_map.get(normalize_header(alias))
        if k is not None:
            return row.get(k)
    return None


def to_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def to_int(v: Any) -> Optional[int]:
    s = to_str(v)
    if s == "":
        return None
    try:
        return int(s)
    except Exception:
        return None


def to_decimal(v: Any, *, source: str) -> Optional[float]:
    s = to_str(v)
    if s == "":
        return None
    try:
        return parse_decimal(s, source=source)
    except Exception:
        return None


# =========================
# 2.3 Dataset entrypoint (calls per-file validators)
# =========================


def validate_dataset(dataset_dir: Path) -> ValidationResult:
    """
    Validates an entire dataset directory (typically data/active or a staging dataset folder).
    Expects canonical filenames (2.1):
      - articles.csv
      - tiers.csv
      - supplier_factors.csv
      - transport.csv
      - customers.xlsx
    """
    errors: list[ValidationError] = []
    warnings: list[ValidationWarning] = []

    def merge(res: ValidationResult) -> None:
        errors.extend(res.errors)
        warnings.extend(res.warnings)

    # Local imports to avoid circular imports at module import time
    from .articles import validate_articles_csv
    from .tiers import validate_tiers_csv
    from .supplier_factors import validate_supplier_factors_csv
    from .transport import validate_transport_csv
    from .customers import validate_customers_xlsx

    try:
        merge(validate_articles_csv(dataset_dir / "articles.csv"))
    except ParseError as e:
        errors.append(_err("articles", None, None, "PARSE_ERROR", str(e)))

    try:
        merge(validate_tiers_csv(dataset_dir / "tiers.csv"))
    except ParseError as e:
        errors.append(_err("tiers", None, None, "PARSE_ERROR", str(e)))

    try:
        merge(validate_supplier_factors_csv(dataset_dir / "supplier_factors.csv"))
    except ParseError as e:
        errors.append(_err("supplier_factors", None, None, "PARSE_ERROR", str(e)))

    try:
        merge(validate_transport_csv(dataset_dir / "transport.csv"))
    except ParseError as e:
        errors.append(_err("transport", None, None, "PARSE_ERROR", str(e)))

    try:
        merge(validate_customers_xlsx(dataset_dir / "customers.xlsx"))
    except ParseError as e:
        errors.append(_err("customers", None, None, "PARSE_ERROR", str(e)))

    return ValidationResult(ok=len(errors) == 0, errors=errors, warnings=warnings)
